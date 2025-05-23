from sqlalchemy.orm import Session
from app.models.registered_vehicle import RegisteredVehicle
from app.schemas.registered_vehicle import VehicleCreate, VehicleResponse
import re
from fastapi import HTTPException, UploadFile
from pathlib import Path
import shutil
import os
from openpyxl import load_workbook
from io import BytesIO
class VehicleService:

    @staticmethod
    def clean_license_plate(license_plate: str) -> str:
        """
        Xóa các ký tự đặc biệt khỏi biển số xe, chỉ giữ lại chữ và số.

        Args:
            license_plate: Biển số xe gốc

        Returns:
            str: Biển số xe đã được làm sạch
        """
        return re.sub(r'[^a-zA-Z0-9]', '', license_plate)

    @staticmethod
    def create_vehicle(db: Session, vehicle: VehicleCreate, image: UploadFile) -> VehicleResponse:
        """
        Tạo một phương tiện mới trong cơ sở dữ liệu.

        Args:
            db: SQLAlchemy session
            vehicle: Schema chứa thông tin phương tiện

        Returns:
            VehicleResponse: Schema chứa thông tin phương tiện đã tạo

        Raises:
            ValueError: Nếu biển số đã tồn tại
        """
        clean_plate = VehicleService.clean_license_plate(vehicle.license_plate)
        vehicle.license_plate = clean_plate

        # Kiểm tra license_plate đã tồn tại chưa
        db_vehicle = db.query(RegisteredVehicle).filter(RegisteredVehicle.license_plate == vehicle.license_plate).first()
        if db_vehicle:
            raise ValueError("License plate already registered")
        
        # Kiểm tra định dạng file hình ảnh
        allowed_extensions = {'.jpg', '.jpeg', '.png'}
        file_extension = Path(image.filename).suffix.lower()
        if file_extension not in allowed_extensions:
            raise ValueError("Invalid image format. Only JPG, JPEG, PNG are allowed")

        # Tạo thư mục lưu trữ nếu chưa tồn tại
        upload_dir = Path("uploads/vehicles")
        upload_dir.mkdir(parents=True, exist_ok=True)
        
        file_name = f"{vehicle.license_plate}{file_extension}"
        file_path = upload_dir / file_name

        # Lưu file hình ảnh
        with file_path.open("wb") as buffer:
            shutil.copyfileobj(image.file, buffer)

        # Tạo đường dẫn tương đối để lưu vào database
        relative_path = f"/uploads/vehicles/{file_name}"

        # Tạo bản ghi phương tiện
        db_vehicle = RegisteredVehicle(
            license_plate=vehicle.license_plate,
            owner_name=vehicle.owner_name,
            phone_number=vehicle.phone_number,
            company=vehicle.company,
            floor_number=vehicle.floor_number,
            image_path=relative_path
        )
        db.add(db_vehicle)
        db.commit()
        db.refresh(db_vehicle)
        return VehicleResponse.from_orm(db_vehicle)
    
    @staticmethod
    def import_vehicles_excel(contents: bytes, db: Session) -> VehicleResponse:
        workbook = load_workbook(filename=BytesIO(contents))
        sheet = workbook.active
        added_vehicles = []
        base_image_dir = "uploads/vehicles"  # Thư mục chứa ảnh trên server

        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            license_plate, owner_name, phone_number, company, floor_number, image_path = row[:6]

            if not license_plate:
                continue

            # Kiểm tra xem phương tiện đã tồn tại
            exists = db.query(RegisteredVehicle).filter_by(license_plate=license_plate).first()
            if exists:
                continue

            # Tạo image_path đầy đủ từ license_plate
            if not image_path:
                raise HTTPException(status_code=400, detail=f"Dòng {i}: image_path không được để trống")

            # Giả sử image_path trong Excel chỉ là license_plate
            if image_path != license_plate:
                raise HTTPException(status_code=400, detail=f"Dòng {i}: image_path '{image_path}' không khớp với biển số '{license_plate}'")

            # Tạo danh sách các file ảnh có thể có với các đuôi khác nhau
            allowed_extensions = ['.jpg', '.jpeg', '.png']
            image_full_path = None
            image_filename = None
            for ext in allowed_extensions:
                possible_filename = f"{license_plate}{ext}"
                possible_full_path = os.path.join(base_image_dir, possible_filename)
                if os.path.exists(possible_full_path):
                    image_filename = possible_filename
                    image_full_path = possible_full_path
                    break

            if not image_full_path:
                raise HTTPException(status_code=400, detail=f"Dòng {i}: Không tìm thấy file ảnh cho biển số '{license_plate}' trong '{base_image_dir}'")

            # Tạo image_path đầy đủ để lưu vào DB
            full_image_path = os.path.join("uploads/vehicles", image_filename).replace("\\", "/")

            vehicle = RegisteredVehicle(
                license_plate=license_plate,
                owner_name=owner_name,
                phone_number=phone_number,
                company=company,
                floor_number=floor_number,
                image_path=full_image_path
            )
            db.add(vehicle)
            added_vehicles.append(1)

        if not added_vehicles:
            raise HTTPException(status_code=400, detail="Không có phương tiện nào để thêm")

        db.commit()
        # Lấy phương tiện cuối cùng để trả về (hoặc điều chỉnh theo logic của bạn)
        db.refresh(vehicle)
        return VehicleResponse.from_orm(vehicle)

    @staticmethod
    def get_vehicle_by_license_plate(db: Session, license_plate: str) -> VehicleResponse:
        """
        Lấy thông tin phương tiện theo biển số.

        Args:
            db: SQLAlchemy session
            license_plate: Biển số phương tiện

        Returns:
            VehicleResponse: Schema chứa thông tin phương tiện

        Raises:
            ValueError: Nếu phương tiện không tồn tại
        """
        clean_plate = VehicleService.clean_license_plate(license_plate)
        license_plate = clean_plate

        db_vehicle = db.query(RegisteredVehicle).filter(RegisteredVehicle.license_plate == license_plate).first()
        if not db_vehicle:
            raise ValueError("Vehicle not found")
        return VehicleResponse.from_orm(db_vehicle)
    
    @staticmethod
    def get_all_vehicles(db: Session) -> VehicleResponse:
        """
        Lấy thông tin phương tiện theo biển số.

        Args:
            db: SQLAlchemy session
            license_plate: Biển số phương tiện

        Returns:
            VehicleResponse: Schema chứa thông tin phương tiện

        Raises:
            ValueError: Nếu phương tiện không tồn tại
        """
        vehicles = db.query(RegisteredVehicle).all()
        return [VehicleResponse.from_orm(vehicle) for vehicle in vehicles]

    @staticmethod
    def update_vehicle(db: Session, license_plate: str, vehicle_update: VehicleCreate) -> VehicleResponse:
        """
        Cập nhật thông tin phương tiện theo biển số.

        Args:
            db: SQLAlchemy session
            license_plate: Biển số phương tiện cần cập nhật
            vehicle_update: Schema chứa thông tin cập nhật

        Returns:
            VehicleResponse: Schema chứa thông tin phương tiện sau khi cập nhật

        Raises:
            ValueError: Nếu phương tiện không tồn tại
        """

        clean_plate = VehicleService.clean_license_plate(license_plate)
        license_plate = clean_plate

        clean_plate = VehicleService.clean_license_plate(vehicle_update.license_plate)
        vehicle_update.license_plate = clean_plate

        db_vehicle = db.query(RegisteredVehicle).filter(RegisteredVehicle.license_plate == license_plate).first()
        if not db_vehicle:
            raise ValueError("Vehicle not found")

        # Cập nhật các trường từ vehicle_update
        for key, value in vehicle_update.dict(exclude_unset=True).items():
            setattr(db_vehicle, key, value)

        db.commit()
        db.refresh(db_vehicle)
        return VehicleResponse.from_orm(db_vehicle)

    @staticmethod
    def delete_vehicle(db: Session, license_plate: str) -> None:
        """
        Xóa phương tiện theo biển số.

        Args:
            db: SQLAlchemy session
            license_plate: Biển số phương tiện cần xóa

        Raises:
            ValueError: Nếu phương tiện không tồn tại
        """

        clean_plate = VehicleService.clean_license_plate(license_plate)
        license_plate = clean_plate

        db_vehicle = db.query(RegisteredVehicle).filter(RegisteredVehicle.license_plate == license_plate).first()
        if not db_vehicle:
            raise ValueError("Vehicle not found")

        if db_vehicle.image_path:
            BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../'))
            full_image_path = os.path.join(BASE_DIR, db_vehicle.image_path.lstrip("/").replace("/", os.sep))
            if os.path.exists(full_image_path):
                os.remove(full_image_path)
                print(f"Đã xóa file ảnh: {full_image_path}")
            else:
                print(f"File ảnh không tồn tại: {full_image_path}")

        db.delete(db_vehicle)
        db.commit()

    